import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class EDIDelforParser:
    def __init__(self, filepath=None):
        self.root = tk.Tk()
        self.root.title("EDI MINEBEA Parser")
        self.root.geometry("1200x800")
        
        # Hlavní data
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        
        # Handle window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.setup_ui()
        
        # If filepath was provided, load it automatically
        if filepath:
            self.load_file(filepath)
        
    def setup_ui(self):
        # Hlavní frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tlačítka pro ovládání
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(btn_frame, text="Export do Excelu", command=self.export_to_excel).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Zpět na hlavní okno", command=self.back_to_main).pack(side=tk.LEFT, padx=(10, 0))
        
        # Notebook pro záložky
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Záložka - Základní informace
        self.info_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.info_frame, text="Základní informace")
        
        # Záložka - Dodávky
        self.delivery_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.delivery_frame, text="Plán dodávek")
        
        # Záložka - Statistiky
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="Statistiky")
        
        self.setup_info_tab()
        self.setup_delivery_tab()
        self.setup_stats_tab()
        
    def setup_info_tab(self):
        # Scrollable text widget pro základní informace
        text_frame = ttk.Frame(self.info_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.info_text = tk.Text(text_frame, wrap=tk.WORD, font=('Courier', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar.set)
        
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def get_scc_description(self, scc_code):
        """Convert SCC code to descriptive name"""
        scc_mapping = {
            '10': 'Backlog',
            '1': 'FIX',
            '4': 'Forecast',
            '': 'Neznámé',
        }
        return scc_mapping.get(scc_code, f'Neznámý kód: {scc_code}')
        
    def setup_delivery_tab(self):
        # Treeview pro plán dodávek
        tree_frame = ttk.Frame(self.delivery_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ('Datum od', 'Datum do', 'Množství', 'Typ', 'SCC')
        self.delivery_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        # Definice sloupců
        for col in columns:
            self.delivery_tree.heading(col, text=col)
            self.delivery_tree.column(col, width=120)
        
        # Scrollbary
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.delivery_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.delivery_tree.xview)
        self.delivery_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.delivery_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
    def setup_stats_tab(self):
        # Statistiky
        stats_frame = ttk.Frame(self.stats_frame)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.stats_text = tk.Text(stats_frame, wrap=tk.WORD, font=('Courier', 10))
        stats_scrollbar = ttk.Scrollbar(stats_frame, orient=tk.VERTICAL, command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scrollbar.set)
        
        self.stats_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        stats_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def parse_date(self, date_str, format_code):
        """Parsuje datum podle EDI formátu"""
        try:
            if format_code == '203':  # CCYYMMDDHHMMSS
                return datetime.strptime(date_str, '%Y%m%d%H%M%S').strftime('%d.%m.%Y %H:%M:%S')
            elif format_code == '102':  # CCYYMMDD
                return datetime.strptime(date_str, '%Y%m%d').strftime('%d.%m.%Y')
            else:
                return date_str
        except:
            return date_str
    
    def parse_edi_datetime(self, datetime_str):
        """Parsuje EDI datum/čas z UNB segmentu (YYMMDD:HHMM)"""
        try:
            if ':' in datetime_str:
                date_part, time_part = datetime_str.split(':')
                # Přidáme 20 na začátek roku (předpokládáme 21. století)
                full_date = '20' + date_part
                formatted_date = datetime.strptime(full_date, '%Y%m%d').strftime('%d.%m.%Y')
                formatted_time = datetime.strptime(time_part, '%H%M').strftime('%H:%M')
                return f"{formatted_date} {formatted_time}"
            return datetime_str
        except:
            return datetime_str
    
    def parse_edi_file(self, content):
        """Parsuje EDI DELFOR soubor"""
        lines = content.strip().split("'")
        
        # Reset dat
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        
        current_delivery = {}
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # UNB - Interchange header
            if line.startswith('UNB'):
                parts = line.split('+')
                if len(parts) >= 5:
                    self.header_info['Odesílatel'] = parts[2]
                    # Uložíme kód příjemce, název doplníme později z NAD segmentu
                    self.header_info['Příjemce_kód'] = parts[3]
                    self.header_info['Datum/Čas'] = self.parse_edi_datetime(parts[4])
                    
            # BGM - Beginning of message
            elif line.startswith('BGM'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Číslo zprávy'] = parts[2]
                    
            # DTM - Date/time
            elif line.startswith('DTM'):
                parts = line.split('+')
                if len(parts) >= 2:
                    dtm_parts = parts[1].split(':')
                    if len(dtm_parts) >= 3:
                        date_formatted = self.parse_date(dtm_parts[1], dtm_parts[2])
                        if dtm_parts[0] == '137':
                            self.header_info['Datum dokumentu'] = date_formatted
                        elif dtm_parts[0] == '63':
                            current_delivery['Datum do'] = date_formatted
                        elif dtm_parts[0] == '64':
                            current_delivery['Datum od'] = date_formatted
                            
            # NAD - Name and address
            elif line.startswith('NAD'):
                parts = line.split('+')
                if len(parts) >= 3:
                    role = parts[1]
                    code = parts[2] if len(parts) > 2 else ''
                    
                    # Debug - vypíšeme co parsujeme
                    print(f"NAD Debug - Role: {role}, Code: {code}, Parts: {parts}")
                    
                    # Název společnosti je v parts[4] (index 4)
                    name = parts[4] if len(parts) > 4 else ''
                    
                    # Adresa začíná od parts[5]
                    address_parts = []
                    for i in range(5, len(parts)):
                        if parts[i]:  # Přidáme pouze neprázdné části
                            address_parts.append(parts[i])
                    
                    full_address = ', '.join(address_parts) if address_parts else ''
                    
                    if role == 'BY':
                        self.partner_info['Kupující'] = name
                        if full_address:
                            self.partner_info['Kupující'] += f", {full_address}"
                    elif role == 'SE':
                        # Zkontrolujeme, zda SE obsahuje kód příjemce z UNB
                        print(f"SE Debug - Checking code: {code}, name: {name}")
                        if '1000500120' in code:
                            print(f"Found matching code! Setting recipient to: {name}")
                            self.header_info['Příjemce'] = name
                        
                        # Pro prodávajícího použijeme název + adresu
                        if full_address:
                            self.partner_info['Prodávající'] = f"{name}, {full_address}"
                        else:
                            self.partner_info['Prodávající'] = name
                    elif role == 'CN':
                        if full_address:
                            self.partner_info['Dodací adresa'] = f"{name}, {full_address}"
                        else:
                            self.partner_info['Dodací adresa'] = name
                        
            # LIN - Line item
            elif line.startswith('LIN'):
                parts = line.split('+')
                if len(parts) >= 4:
                    self.header_info['Číslo položky'] = parts[3]
                    
            # PIA - Product identification
            elif line.startswith('PIA'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Kód produktu'] = parts[2]
                    
            # QTY - Quantity
            elif line.startswith('QTY'):
                parts = line.split('+')
                if len(parts) >= 2:
                    qty_parts = parts[1].split(':')
                    if len(qty_parts) >= 3:
                        qty_type = qty_parts[0]
                        quantity = qty_parts[1]
                        unit = qty_parts[2]
                        
                        if qty_type == '113':  # Cumulative quantity
                            current_delivery['Množství'] = quantity
                            current_delivery['Jednotka'] = unit
                            current_delivery['Typ'] = 'Kumulativní'
                        elif qty_type == '70':  # Minimum quantity
                            current_delivery['Množství'] = quantity
                            current_delivery['Jednotka'] = unit
                            current_delivery['Typ'] = 'Minimální'
                        elif qty_type == '78':  # Maximum quantity
                            current_delivery['Množství'] = quantity
                            current_delivery['Jednotka'] = unit
                            current_delivery['Typ'] = 'Maximální'
                            
            # SCC - Scheduling conditions
            elif line.startswith('SCC'):
                parts = line.split('+')
                if len(parts) >= 2:
                    current_delivery['SCC'] = parts[1]
                    
                    # Pokud máme kompletní dodávku, přidáme ji
                    if 'Datum od' in current_delivery and 'Množství' in current_delivery:
                        self.delivery_schedules.append(current_delivery.copy())
                        current_delivery = {'SCC': parts[1]}  # Zachováme SCC pro další dodávky
    
    def load_file(self, filepath):
        """Načte EDI soubor"""
        try:
            # Check if the window still exists
            if not hasattr(self, 'root') or not self.root.winfo_exists():
                return False
                
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
                
            self.parse_edi_file(content)
            
            # Check again before updating UI
            if hasattr(self, 'root') and self.root.winfo_exists():
                self.display_data()
                return True
            return False
            
        except Exception as e:
            # Safely show error if window still exists
            if hasattr(self, 'root') and self.root.winfo_exists():
                messagebox.showerror("Chyba", f"Nelze načíst soubor: {str(e)}")
            return False
    
    def display_data(self):
        """Zobrazí naparsovaná data"""
        # Check if window still exists
        if not hasattr(self, 'info_text') or not hasattr(self, 'root') or not self.root.winfo_exists():
            return
            
        try:
            # Základní informace
            self.info_text.delete(1.0, tk.END)
            info_content = "=== HLAVIČKA DOKUMENTU ===\n"
            for key, value in self.header_info.items():
                # Přeskočíme pomocný klíč
                if key != 'Příjemce_kód':
                    info_content += f"{key}: {value}\n"
            
            # Pokud nemáme název příjemce, zobrazíme alespoň kód
            if 'Příjemce' not in self.header_info and 'Příjemce_kód' in self.header_info:
                info_content += f"Příjemce: {self.header_info['Příjemce_kód']}\n"
        except Exception as e:
            # Skip if there's an error during display
            print(f"Error displaying data: {e}")
            return
        
        info_content += "\n=== INFORMACE O PARTNERECH ===\n"
        for key, value in self.partner_info.items():
            info_content += f"{key}: {value}\n"
        
        self.info_text.insert(1.0, info_content)
        
        # Plán dodávek
        for item in self.delivery_tree.get_children():
            self.delivery_tree.delete(item)
            
        for delivery in self.delivery_schedules:
            scc_code = delivery.get('SCC', '')
            scc_desc = self.get_scc_description(scc_code)
            self.delivery_tree.insert('', tk.END, values=(
                delivery.get('Datum od', ''),
                delivery.get('Datum do', ''),
                delivery.get('Množství', ''),
                delivery.get('Typ', ''),
                scc_desc
            ))
        
        # Statistiky
        self.stats_text.delete(1.0, tk.END)
        stats_content = "=== STATISTIKY ===\n"
        stats_content += f"Celkový počet dodávek: {len(self.delivery_schedules)}\n"
        
        total_qty = sum(int(d.get('Množství', 0)) for d in self.delivery_schedules if d.get('Množství', '').isdigit())
        stats_content += f"Celkové množství: {total_qty:,} kusů\n"
        
        # Statistiky podle typu
        type_stats = {}
        for delivery in self.delivery_schedules:
            delivery_type = delivery.get('Typ', 'Neznámý')
            if delivery_type not in type_stats:
                type_stats[delivery_type] = {'počet': 0, 'množství': 0}
            type_stats[delivery_type]['počet'] += 1
            if delivery.get('Množství', '').isdigit():
                type_stats[delivery_type]['množství'] += int(delivery.get('Množství', 0))
        
        stats_content += "\n=== STATISTIKY PODLE TYPU ===\n"
        for delivery_type, stats in type_stats.items():
            stats_content += f"{delivery_type}: {stats['počet']} dodávek, {stats['množství']:,} kusů\n"
        
        self.stats_text.insert(1.0, stats_content)
    
    def get_week_number(self, date_str):
        """Převede řetězec s datem na číslo kalendářního týdne (WW)"""
        try:
            day, month, year = map(int, date_str.split('.'))
            dt = date(year, month, day)
            return dt.isocalendar()[1]
        except (ValueError, AttributeError):
            return ""

    def export_to_excel(self):
        """Exportuje data o dodávkách do Excelu s kalendářními týdny"""
        if not self.delivery_schedules:
            messagebox.showwarning("Upozornění", "Žádná data k exportu")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dodávky"

            # Hlavičky
            headers = ["Týden", "Datum od", "Datum do", "Množství", "Typ", "SCC"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Data
            row_num = 2
            for delivery in self.delivery_schedules:
                date_from = delivery.get('Datum od', '')
                week_num = self.get_week_number(date_from) if date_from else ""
                scc_code = delivery.get('SCC', '')
                scc_desc = self.get_scc_description(scc_code)
                
                ws.cell(row=row_num, column=1, value=week_num)
                ws.cell(row=row_num, column=2, value=date_from)
                ws.cell(row=row_num, column=3, value=delivery.get('Datum do', ''))
                ws.cell(row=row_num, column=4, value=delivery.get('Množství', ''))
                ws.cell(row=row_num, column=5, value=delivery.get('Typ', ''))
                ws.cell(row=row_num, column=6, value=scc_desc)
                row_num += 1

            # Automatické přizpůsobení šířky sloupců
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 30)

            # Uložení souboru
            filename = f"dodavky_minebea_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Hotovo", f"Data byla úspěšně exportována do souboru:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Chyba", f"Chyba při exportu do Excelu: {str(e)}")

    def on_closing(self):
        """Handle window close event"""
        self.root.destroy()  # Close the current window
        
    def back_to_main(self):
        """Closes the current window"""
        self.root.destroy()
    
    def run(self):
        """Spustí aplikaci"""
        self.root.mainloop()

if __name__ == "__main__":
    app = EDIDelforParser()
    app.run()