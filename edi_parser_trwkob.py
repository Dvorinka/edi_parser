import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class EDITrwkobParser:
    def __init__(self, filepath=None):
        self.root = tk.Tk()
        self.root.title("EDI TRWKOB Parser")
        self.root.geometry("1200x800")
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        self.setup_ui()
        self.main_window = None

    def setup_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(btn_frame, text="Export do Excelu", command=self.export_to_excel).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Zpět na hlavní okno", command=self.back_to_main).pack(side=tk.LEFT, padx=(10, 0))
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.info_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.info_frame, text="Základní informace")
        self.delivery_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.delivery_frame, text="Plán dodávek")
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="Statistiky")
        self.setup_info_tab()
        self.setup_delivery_tab()
        self.setup_stats_tab()

    def setup_info_tab(self):
        text_frame = ttk.Frame(self.info_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.info_text = tk.Text(text_frame, wrap=tk.WORD, font=('Courier', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar.set)
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_delivery_tab(self):
        tree_frame = ttk.Frame(self.delivery_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        columns = ('Datum od', 'Datum do', 'Množství', 'Typ', 'SCC')
        self.delivery_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.delivery_tree.heading(col, text=col)
            self.delivery_tree.column(col, width=120)
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.delivery_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.delivery_tree.xview)
        self.delivery_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.delivery_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    def setup_stats_tab(self):
        stats_frame = ttk.Frame(self.stats_frame)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.stats_text = tk.Text(stats_frame, wrap=tk.WORD, font=('Courier', 10))
        stats_scrollbar = ttk.Scrollbar(stats_frame, orient=tk.VERTICAL, command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scrollbar.set)
        self.stats_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        stats_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def parse_date(self, date_str, format_code):
        try:
            if format_code == '102':
                return datetime.strptime(date_str, '%Y%m%d').strftime('%d.%m.%Y')
            else:
                return date_str
        except:
            return date_str

    def parse_edi_datetime(self, datetime_str):
        """Parsuje EDI datum/čas z UNB segmentu (YYMMDD:HHMM)"""
        try:
            if ':' in datetime_str:
                date_part, time_part = datetime_str.split(':')
                full_date = '20' + date_part
                formatted_date = datetime.strptime(full_date, '%Y%m%d').strftime('%d.%m.%Y')
                formatted_time = datetime.strptime(time_part, '%H%M').strftime('%H:%M')
                return f"{formatted_date} {formatted_time}"
            return datetime_str
        except:
            return datetime_str

    def load_file(self, filepath):
        """Load and parse the specified EDI file"""
        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            self.parse_edi_file(content)
            self.display_data()
            return True
        except Exception as e:
            messagebox.showerror("Chyba", f"Nelze načíst soubor: {str(e)}")
            return False

    def parse_edi_file(self, content):
        lines = content.strip().split("'")
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        current_delivery = {}
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith('UNB'):
                parts = line.split('+')
                if len(parts) >= 5:
                    self.header_info['Odesílatel'] = parts[2]
                    self.header_info['Příjemce_kód'] = parts[3]
                    self.header_info['Datum/Čas'] = self.parse_edi_datetime(parts[4])
            elif line.startswith('BGM'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Číslo zprávy'] = parts[2]
            elif line.startswith('DTM'):
                parts = line.split('+')
                if len(parts) >= 2:
                    dtm_parts = parts[1].split(':')
                    if len(dtm_parts) >= 2:
                        date_formatted = self.parse_date(dtm_parts[1], dtm_parts[2] if len(dtm_parts) > 2 else '')
                        if dtm_parts[0] == '137':
                            self.header_info['Datum dokumentu'] = date_formatted
                        elif dtm_parts[0] == '63':
                            current_delivery['Datum do'] = date_formatted
                        elif dtm_parts[0] == '64':
                            current_delivery['Datum od'] = date_formatted
            elif line.startswith('NAD'):
                parts = line.split('+')
                if len(parts) >= 3:
                    role = parts[1]
                    code = parts[2]
                    name = parts[4] if len(parts) > 4 else ''
                    address_parts = []
                    for i in range(5, len(parts)):
                        if parts[i]:
                            address_parts.append(parts[i])
                    full_address = ', '.join(address_parts) if address_parts else ''
                    if role == 'BY':
                        # Kupující: prefer name, fallback to code
                        self.partner_info['Kupující'] = name if name else code
                        if full_address:
                            self.partner_info['Kupující'] += f", {full_address}"
                    elif role == 'SE':
                        # Prodávající: prefer name, fallback to code
                        self.header_info['Příjemce'] = name if name else code
                        if full_address:
                            self.partner_info['Prodávající'] = f"{name if name else code}, {full_address}"
                        else:
                            self.partner_info['Prodávající'] = name if name else code
                    elif role == 'CN':
                        # Dodací adresa: show code if name missing
                        if full_address:
                            self.partner_info['Dodací adresa'] = f"{name if name else code}, {full_address}"
                        else:
                            self.partner_info['Dodací adresa'] = name if name else code
            elif line.startswith('LIN'):
                parts = line.split('+')
                if len(parts) >= 4:
                    self.header_info['Číslo položky'] = parts[3]
            elif line.startswith('PIA'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Kód produktu'] = parts[2]
            elif line.startswith('QTY'):
                parts = line.split('+')
                if len(parts) >= 2:
                    qty_parts = parts[1].split(':')
                    if len(qty_parts) >= 3:
                        qty_type = qty_parts[0]
                        quantity = qty_parts[1]
                        unit = qty_parts[2]
                        if qty_type == '113':
                            current_delivery['Množství'] = quantity
                            current_delivery['Jednotka'] = unit
                            current_delivery['Typ'] = 'Kumulativní'
                        elif qty_type == '70':
                            current_delivery['Množství'] = quantity
                            current_delivery['Jednotka'] = unit
                            current_delivery['Typ'] = 'Minimální'
                        elif qty_type == '78':
                            current_delivery['Množství'] = quantity
                            current_delivery['Jednotka'] = unit
                            current_delivery['Typ'] = 'Maximální'
            elif line.startswith('SCC'):
                parts = line.split('+')
                if len(parts) >= 2:
                    current_delivery['SCC'] = parts[1]
                    if 'Datum od' in current_delivery and 'Množství' in current_delivery:
                        self.delivery_schedules.append(current_delivery.copy())
                        current_delivery = {'SCC': parts[1]}

    def display_data(self):
        self.info_text.delete(1.0, tk.END)
        info_content = "=== HLAVIČKA DOKUMENTU ===\n"
        for key, value in self.header_info.items():
            if key != 'Příjemce_kód':
                info_content += f"{key}: {value}\n"
        if 'Příjemce' not in self.header_info and 'Příjemce_kód' in self.header_info:
            info_content += f"Příjemce: {self.header_info['Příjemce_kód']}\n"
        info_content += "\n=== INFORMACE O PARTNERECH ===\n"
        for key, value in self.partner_info.items():
            info_content += f"{key}: {value}\n"
        self.info_text.insert(1.0, info_content)
        for item in self.delivery_tree.get_children():
            self.delivery_tree.delete(item)
        for delivery in self.delivery_schedules:
            scc_code = delivery.get('SCC', '')
            scc_desc = self.get_scc_description(scc_code)
            self.delivery_tree.insert('', tk.END, values=(
                delivery.get('Datum od', ''),
                delivery.get('Datum do', ''),
                delivery.get('Množství', ''),
                delivery.get('Typ', ''),
                scc_desc
            ))
        self.stats_text.delete(1.0, tk.END)
        stats_content = "=== STATISTIKY ===\n"
        stats_content += f"Celkový počet dodávek: {len(self.delivery_schedules)}\n"
        total_qty = sum(int(d.get('Množství', 0)) for d in self.delivery_schedules if d.get('Množství', '').isdigit())
        stats_content += f"Celkové množství: {total_qty:,} kusů\n"
        type_stats = {}
        for delivery in self.delivery_schedules:
            delivery_type = delivery.get('Typ', 'Neznámý')
            if delivery_type not in type_stats:
                type_stats[delivery_type] = {'počet': 0, 'množství': 0}
            type_stats[delivery_type]['počet'] += 1
            if delivery.get('Množství', '').isdigit():
                type_stats[delivery_type]['množství'] += int(delivery.get('Množství', 0))
        stats_content += "\n=== STATISTIKY PODLE TYPU ===\n"
        for delivery_type, stats in type_stats.items():
            stats_content += f"{delivery_type}: {stats['počet']} dodávek, {stats['množství']:,} kusů\n"
        self.stats_text.insert(1.0, stats_content)

    def get_week_number(self, date_str):
        """Convert date string to ISO week number (WW)"""
        try:
            day, month, year = map(int, date_str.split('.'))
            dt = date(year, month, day)
            return dt.isocalendar()[1]
        except (ValueError, AttributeError):
            return ""
            
    def get_scc_description(self, scc_code):
        """Convert SCC code to descriptive name"""
        scc_mapping = {
            '10': 'Backlog',
            '1': 'FIX',
            '4': 'Forecast',
            '': 'Neznámé',
        }
        return scc_mapping.get(scc_code, f'Neznámý kód: {scc_code}')

    def export_to_excel(self):
        """Export delivery data to Excel with calendar weeks"""
        if not self.delivery_schedules:
            messagebox.showwarning("Upozornění", "Žádná data k exportu")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dodávky"

            # Headers
            headers = ["Týden", "Datum", "Množství", "Typ", "SCC"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Data
            row_num = 2
            for delivery in self.delivery_schedules:
                date_str = delivery.get('Datum od', '')
                week_num = self.get_week_number(date_str) if date_str else ""
                scc_code = delivery.get('SCC', '')
                scc_desc = self.get_scc_description(scc_code)
                
                ws.cell(row=row_num, column=1, value=week_num)
                ws.cell(row=row_num, column=2, value=date_str)
                ws.cell(row=row_num, column=3, value=delivery.get('Množství', ''))
                ws.cell(row=row_num, column=4, value=delivery.get('Typ', ''))
                ws.cell(row=row_num, column=5, value=scc_desc)
                row_num += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 30)

            # Save the file
            filename = f"dodavky_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Hotovo", f"Data byla úspěšně exportována do souboru:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Chyba", f"Chyba při exportu do Excelu: {str(e)}")

    def back_to_main(self):
        """Closes the current window and returns to the main application"""
        # Close current window
        self.root.destroy()
        # Return to main window
        if self.main_window:
            self.main_window.root.deiconify()  # Show the main window if it exists

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = EDITrwkobParser()
    app.run()
