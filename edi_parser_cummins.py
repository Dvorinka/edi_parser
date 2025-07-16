import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class EDIDelforCumminsParser:
    def __init__(self, filepath=None):
        self.root = tk.Tk()
        self.root.title("EDI Cummins Parser")
        self.root.geometry("1200x800")
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        self.line_items = []
        
        # Handle window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.setup_ui()
        if filepath:
            self.load_file(filepath)

    def setup_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        # Add buttons
        ttk.Button(btn_frame, text="Zpět na hlavní okno", command=self.back_to_main).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(btn_frame, text="Export do Excelu", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.info_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.info_frame, text="Základní informace")
        self.delivery_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.delivery_frame, text="Plán dodávek")
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="Statistiky")
        self.setup_info_tab()
        self.setup_delivery_tab()
        self.setup_stats_tab()

    def setup_info_tab(self):
        text_frame = ttk.Frame(self.info_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.info_text = tk.Text(text_frame, wrap=tk.WORD, font=('Courier', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar.set)
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_delivery_tab(self):
        tree_frame = ttk.Frame(self.delivery_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # Removed 'Jednotka' column as requested
        columns = ('Položka', 'Popis', 'Datum', 'Množství', 'Typ', 'SCC', 'Release')
        self.delivery_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.delivery_tree.heading(col, text=col)
            if col == 'Popis':
                self.delivery_tree.column(col, width=200)
            elif col == 'Položka':
                self.delivery_tree.column(col, width=100)
            else:
                self.delivery_tree.column(col, width=80)
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.delivery_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.delivery_tree.xview)
        self.delivery_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.delivery_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    def setup_stats_tab(self):
        stats_frame = ttk.Frame(self.stats_frame)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.stats_text = tk.Text(stats_frame, wrap=tk.WORD, font=('Courier', 10))
        stats_scrollbar = ttk.Scrollbar(stats_frame, orient=tk.VERTICAL, command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scrollbar.set)
        self.stats_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        stats_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def parse_date(self, date_str, format_code):
        try:
            if format_code == '102':
                return datetime.strptime(date_str, '%Y%m%d').strftime('%d.%m.%Y')
            else:
                return date_str
        except:
            return date_str

    def parse_edi_datetime(self, datetime_str):
        try:
            if ':' in datetime_str:
                date_part, time_part = datetime_str.split(':')
                full_date = '20' + date_part
                formatted_date = datetime.strptime(full_date, '%Y%m%d').strftime('%d.%m.%Y')
                formatted_time = datetime.strptime(time_part, '%H%M').strftime('%H:%M')
                return f"{formatted_date} {formatted_time}"
            return datetime_str
        except:
            return datetime_str

    def get_scc_description(self, scc_code):
        scc_map = {
            '10': 'Backlog',
            '1': 'Firm',
            '4': 'Forecast'
        }
        return scc_map.get(scc_code, f'SCC-{scc_code}')

    def parse_edi_file(self, content):
        lines = content.strip().split("'")
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        self.line_items = []
        
        # Current parsing state
        current_part_number = ''
        current_description = ''
        current_location = ''
        current_po = ''
        current_scc = ''
        current_release = ''
        
        # Track current line item details
        current_line_item = None
        
        # Temporary storage for quantity waiting for date
        pending_quantities = []
        
        def create_or_update_line_item():
            nonlocal current_line_item
            if not current_part_number:
                return None
                
            line_item = next((item for item in self.line_items 
                           if item['Položka'] == current_part_number), None)
            
            if not line_item:
                line_item = {
                    'Položka': current_part_number,
                    'Popis': current_description,
                    'Objednávka': current_po,
                    'Lokace': current_location,
                    'RFF': {}
                }
                self.line_items.append(line_item)
                
            # Update current line item reference
            current_line_item = line_item
            return line_item

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith('UNB'):
                parts = line.split('+')
                if len(parts) >= 5:
                    self.header_info['Odesílatel'] = parts[2]
                    self.header_info['Příjemce_kód'] = parts[3]
                    self.header_info['Datum/Čas'] = self.parse_edi_datetime(parts[4])

            elif line.startswith('UNH'):
                parts = line.split('+')
                if len(parts) >= 2:
                    self.header_info['ID zprávy'] = parts[1]

            elif line.startswith('BGM'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Číslo zprávy'] = parts[2]

            elif line.startswith('DTM'):
                parts = line.split('+')
                if len(parts) >= 2:
                    dtm_parts = parts[1].split(':')
                    if len(dtm_parts) >= 3:
                        code = dtm_parts[0]
                        value = dtm_parts[1]
                        fmt = dtm_parts[2]
                        formatted_date = self.parse_date(value, fmt)
                        if code == '137':
                            self.header_info['Datum dokumentu'] = formatted_date
                        elif code == '2':
                            # This is a delivery date - match with pending quantities
                            # Only create entries if we have quantities to process
                            if pending_quantities:
                                # For SCC 10 (Backlog), we only take the first quantity
                                if current_scc == '10' and len(pending_quantities) > 0:
                                    qty_info = pending_quantities[0]
                                    # Create line item if it doesn't exist
                                    line_item = next((item for item in self.line_items if item['Položka'] == current_part_number), None)
                                    if not line_item:
                                        line_item = {
                                            'Položka': current_part_number,
                                            'Popis': current_description,
                                            'Objednávka': current_po,
                                            'Lokace': current_location
                                        }
                                        self.line_items.append(line_item)
                                    
                                    delivery = {
                                        'Položka': current_part_number,
                                        'Popis': current_description,
                                        'Datum': formatted_date,
                                        'Množství': qty_info['quantity'],
                                        'Typ': qty_info['type'],
                                        'SCC': self.get_scc_description(current_scc),
                                        'Release': current_release,
                                        'Objednávka': current_po
                                    }
                                    self.delivery_schedules.append(delivery)
                                else:
                                    # For other SCCs, process all quantities
                                    for qty_info in pending_quantities:
                                        delivery = {
                                            'Položka': current_part_number,
                                            'Popis': current_description,
                                            'Datum': formatted_date,
                                            'Množství': qty_info['quantity'],
                                            'Typ': qty_info['type'],
                                            'SCC': self.get_scc_description(current_scc),
                                            'Release': current_release
                                        }
                                        self.delivery_schedules.append(delivery)
                            pending_quantities.clear()
                            # Don't reset release here to maintain it for next entries

            elif line.startswith('NAD'):
                parts = line.split('+')
                if len(parts) >= 3:
                    role = parts[1]
                    if role in ['SU', 'ST']:
                        name_parts = [p.replace('?+', '').replace('?', '').strip() for p in parts[4:] if p]
                        if role == 'SU':
                            self.partner_info['Dodavatel'] = ' '.join(name_parts)
                        elif role == 'ST':
                            self.partner_info['Příjemce'] = ', '.join(name_parts)

            elif line.startswith('LIN'):
                # Save previous line item if it exists
                if current_part_number:
                    # Process previous line item if exists
                    create_or_update_line_item()
                
                parts = line.split('+')
                # Process LIN segment
                
                if len(parts) >= 4:
                    # Reset part information for new line item
                    current_part_number = ''
                    current_description = ''
                    current_scc = ''
                    current_release = ''
                    current_line_item = None
                    pending_quantities = []
                    
                    # Try to find part number in the LIN segment
                    for i, part in enumerate(parts[3:], 3):  # Skip the first 3 parts (LIN, line number, action code)
                        # Process part
                        if ':' in part:  # If the part contains a colon, it might be a part number
                            part_info = part.split(':')
                            # Process part info
                            if len(part_info) >= 2 and part_info[1] == 'IN':  # Look for part number with 'IN' qualifier
                                current_part_number = part_info[0]
                                # Found part number with IN qualifier
                                break
                            elif not current_part_number:  # If no 'IN' qualifier found, take the first part
                                current_part_number = part_info[0]
                                # Using first part as part number
                    
                    # If still no part number found, try to get it from the last part
                    if not current_part_number and parts[3:]:
                        current_part_number = parts[3].split(':')[0]
                        # Using fallback part number
                    
                    # Final part number processed

            elif line.startswith('IMD'):
                parts = line.split('+')
                if len(parts) >= 4:
                    # Extract item description - fixed to properly handle the format
                    # Looking for the 4th element which contains the description
                    desc_part = parts[3] if len(parts) > 3 else ''
                    
                    # Remove leading colons and extract the actual description
                    if desc_part.startswith(':::'):
                        current_description = desc_part[3:].strip()
                    elif desc_part.startswith('::'):
                        current_description = desc_part[2:].strip()
                    elif desc_part.startswith(':'):
                        current_description = desc_part[1:].strip()
                    else:
                        current_description = desc_part.strip()
                    
                    # Clean up any remaining formatting
                    current_description = current_description.replace(':', '').strip()

            elif line.startswith('LOC'):
                parts = line.split('+')
                if len(parts) >= 3:
                    current_location = parts[2]

            elif line.startswith('RFF'):
                parts = line.split('+')
                # Process RFF segment
                
                if len(parts) >= 2:
                    ref_parts = parts[1].split(':')
                    if len(ref_parts) >= 2:
                        ref_type = ref_parts[0]
                        ref_value = ref_parts[1]
                        
                        # Found RFF reference
                        
                        # Create or update line item if it doesn't exist
                        if not current_line_item:
                            # Create new line item if none exists
                            create_or_update_line_item()
                        
                        # Store the reference in the current line item
                        if current_line_item:
                            if 'RFF' not in current_line_item:
                                current_line_item['RFF'] = {}
                            current_line_item['RFF'][ref_type] = ref_value
                            # RFF stored in line item
                            
                            # Special handling for order numbers
                            if ref_type == 'ON':
                                current_po = ref_value
                                current_line_item['Objednávka'] = current_po
                                # Order number set
                            elif ref_type == 'RE':
                                current_release = ref_value
                                # Clear any pending quantities to ensure release number is applied to new quantities
                                pending_quantities = []
                                # Release number set
                        else:
                            # No line item available for RFF
                            pass

            elif line.startswith('SCC'):
                parts = line.split('+')
                if len(parts) >= 2:
                    current_scc = parts[1]
                    # Clear pending quantities when new SCC starts to prevent duplicates
                    pending_quantities = []
                    # Only reset release for backlog (SCC 10)
                    if current_scc == '10':
                        current_release = ''

            elif line.startswith('QTY'):
                parts = line.split('+')
                if len(parts) >= 2:
                    qty_parts = parts[1].split(':')
                    if len(qty_parts) >= 2:
                        qty_type = qty_parts[0]
                        quantity = qty_parts[1]
                        # Removed unit extraction as we don't need it
                        
                        # Determine quantity type
                        qty_type_desc = 'Neznámý'
                        if qty_type == '1':
                            qty_type_desc = 'Dodávka'
                        elif qty_type == '3':
                            qty_type_desc = 'Kumulativní'
                        elif qty_type == '48':
                            qty_type_desc = 'Plánované'
                        
                        # Store quantity info waiting for corresponding date
                        pending_quantities.append({
                            'quantity': quantity,
                            'type': qty_type_desc
                        })

        # Store line items for reference
        unique_parts = {}
        for delivery in self.delivery_schedules:
            part_num = delivery['Položka']
            if part_num not in unique_parts:
                unique_parts[part_num] = {
                    'Položka': part_num,
                    'Popis': delivery['Popis']
                }
        
        self.line_items = list(unique_parts.values())

    def load_file(self, filepath=None):
        if filepath:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                self.parse_edi_file(content)
                self.display_data()
                return True
            except Exception as e:
                messagebox.showerror("Chyba", f"Nelze načíst soubor: {str(e)}")
                return False
        return False

    def display_data(self):
        # Display header info
        self.info_text.delete(1.0, tk.END)
        info_content = "=== HLAVIČKA DOKUMENTU ===\n"
        for key, value in self.header_info.items():
            if key != 'Příjemce_kód':
                info_content += f"{key}: {value}\n"
        info_content += "\n=== INFORMACE O PARTNERECH ===\n"
        for key, value in self.partner_info.items():
            info_content += f"{key}: {value}\n"
        self.info_text.insert(1.0, info_content)

        # Display delivery schedules
        for item in self.delivery_tree.get_children():
            self.delivery_tree.delete(item)

        # Sort deliveries by date
        def date_sort_key(delivery):
            date_str = delivery.get('Datum', '')
            try:
                return datetime.strptime(date_str, '%d.%m.%Y') if date_str else datetime.max
            except:
                return datetime.max

        sorted_deliveries = sorted(self.delivery_schedules, key=date_sort_key)
        
        for delivery in sorted_deliveries:
            self.delivery_tree.insert('', tk.END, values=(
                delivery.get('Položka', ''),
                delivery.get('Popis', ''),
                delivery.get('Datum', ''),
                delivery.get('Množství', ''),
                delivery.get('Typ', ''),
                delivery.get('SCC', ''),
                delivery.get('Release', '')
            ))

        # Display statistics
        self.stats_text.delete(1.0, tk.END)
        stats_content = "=== STATISTIKY ===\n"
        stats_content += f"Celkový počet dodávek: {len(self.delivery_schedules)}\n"
        stats_content += f"Počet různých položek: {len(self.line_items)}\n"
        
        # Group by SCC
        scc_stats = {}
        total_qty = 0
        for delivery in self.delivery_schedules:
            scc = delivery.get('SCC', 'Neznámý')
            qty_str = delivery.get('Množství', '0')
            try:
                qty = int(qty_str)
                total_qty += qty
                if scc not in scc_stats:
                    scc_stats[scc] = {'count': 0, 'total_qty': 0}
                scc_stats[scc]['count'] += 1
                scc_stats[scc]['total_qty'] += qty
            except:
                pass
        
        stats_content += f"Celkové množství: {total_qty:,} kusů\n\n"
        stats_content += "=== STATISTIKY PO SCC ===\n"
        for scc, stats in scc_stats.items():
            stats_content += f"{scc}: {stats['count']} dodávek, {stats['total_qty']:,} kusů\n"
        
        self.stats_text.insert(1.0, stats_content)

    def on_closing(self):
        """Handle window close event"""
        self.root.destroy()  # Close the current window
        
    def back_to_main(self):
        """Closes the current window"""
        self.root.destroy()

    def get_week_number(self, date_str):
        """Convert date string to ISO week number"""
        try:
            # Handle different date formats
            if '.' in date_str:
                date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
            else:
                date_obj = datetime.strptime(date_str, '%Y%m%d').date()
            return date_obj.isocalendar()[1]  # Returns ISO week number
        except Exception as e:
            print(f"Error parsing date {date_str}: {e}")
            return ""

    def export_to_excel(self):
        """Export delivery data to Excel with calendar weeks"""
        if not self.delivery_schedules:
            messagebox.showwarning("Upozornění", "Žádná data k exportu")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dodávky"

            # Headers with week number
            headers = ["Týden", "Datum", "Položka", "Popis", "Množství", "Typ", "SCC", "Release"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Add data - sort by date from oldest to newest
            def parse_date(date_str):
                try:
                    if '.' in date_str:
                        return datetime.strptime(date_str, '%d.%m.%Y').date()
                    else:
                        return datetime.strptime(date_str, '%Y%m%d').date()
                except:
                    return datetime.min.date()
            
            row_num = 2
            for item in sorted(self.delivery_schedules, key=lambda x: parse_date(x.get('Datum', ''))):
                # Get week number from date
                week_num = self.get_week_number(item.get('Datum', ''))
                
                # Add week number and date in the first two columns
                ws.cell(row=row_num, column=1, value=week_num)
                ws.cell(row=row_num, column=2, value=item.get('Datum', ''))
                
                # Add other data
                ws.cell(row=row_num, column=3, value=item.get('Položka', ''))
                ws.cell(row=row_num, column=4, value=item.get('Popis', ''))
                ws.cell(row=row_num, column=5, value=item.get('Množství', ''))
                ws.cell(row=row_num, column=6, value=item.get('Typ', ''))
                ws.cell(row=row_num, column=7, value=item.get('SCC', ''))
                ws.cell(row=row_num, column=8, value=item.get('Release', ''))
                row_num += 1

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 30)

            # Add a summary sheet with just week and quantity
            ws_summary = wb.create_sheet("Přehled")
            ws_summary.cell(row=1, column=1, value="Týden").font = Font(bold=True)
            ws_summary.cell(row=1, column=2, value="Množství").font = Font(bold=True)
            
            # Group quantities by week (sorted by week number)
            weekly_totals = {}
            for item in sorted(self.delivery_schedules, key=lambda x: parse_date(x.get('Datum', ''))):
                week_num = self.get_week_number(item.get('Datum', ''))
                if week_num:
                    try:
                        qty = int(item.get('Množství', 0))
                        if week_num in weekly_totals:
                            weekly_totals[week_num] += qty
                        else:
                            weekly_totals[week_num] = qty
                    except (ValueError, TypeError):
                        pass
            
            # Add weekly totals to summary
            row = 2
            for week, total in sorted(weekly_totals.items()):
                ws_summary.cell(row=row, column=1, value=week)
                ws_summary.cell(row=row, column=2, value=total)
                row += 1

            # Auto-adjust summary columns
            for col in ws_summary.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_summary.column_dimensions[column].width = min(adjusted_width, 15)

            # Save the file
            filename = f"dodavky_cummins_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Hotovo", f"Data byla úspěšně exportována do souboru:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Chyba", f"Chyba při exportu do Excelu: {str(e)}")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    # When run directly, use the main parser to handle file selection
    from edi_parser_main import EDIUnifiedParser
    EDIUnifiedParser()